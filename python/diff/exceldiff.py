#!/usr/bin/python
#-*- coding:utf-8 -*-

import xlrd
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl import styles


def diff_excel(filepath1,filepath2):
    wb1 = xlrd.open_workbook(filepath1)
    table1=wb1.sheets()[0]
    wb2=xlrd.open_workbook(filepath2)
    table2=wb2.sheets()[0]
    nrows1=table1.nrows
    nrows2=table2.nrows
    print(nrows1,nrows2)
    if nrows1!=nrows2 or table1.ncols!=tables2.ncols:
        print('两个表的行或者列不一致')
        print('excel表的行数为：%d,%d'%(nrows1,nrows2))
        print('excel表的列数为：%d,%d'%(table1.ncols,table2.ncols))
        return   
    for i in range(nrows1):
        row1_value=table1.row_values(i+1)
        row2_value=table2.row_values(i+1)
        if str(row1_value)==str(row2_value):
            pass
        else:
            
            print('第%d行不一致:'%(i+1))
            print(row1_value,row2_value)
            return
            
    
 #以前是不一样输出，现在是不一样在excel中涂黄   
def diff(filepath1,filepath2):
    #flag=True
    yellowfill = styles.PatternFill(fgColor='FFFF00',fill_type='solid') #设置颜色
    wb1=load_workbook(filepath1)
    ws1=wb1.active
    wb2=load_workbook(filepath2)
    #print(wb2.sheetnames)
    ws2=wb2.active
    row1=[]
    row2=[]
    for row in ws1.iter_rows():
        row1.append(row)
    for row in ws2.iter_rows():
        row2.append(row)
    if len(row1)!=len(row2):
        print('数据表行数不一致，分别为：%d,%d'%(len(row1),len(row2)))
        return  
    for i in range(len(row1)-1):
        for j in range(len(row1[0])-1):
            if row1[i][j].value!=row2[i][j].value:
                try:
                    value1=float(row1[i][j].value)
                    value2=float(row2[i][j].value)
                    if abs(value1-value2)>0.01:
                        print('第%d行第%d列数据不一致'%(i,j))
                        print (row1[i][j].value,row2[i][j].value) 
                        ws1.cell(row=i+1,column=j).fill=yellowfill  # 不一样就
                        ws2.cell(row=i+1,column=j).fill=yellowfill   
                except:
                    print('第%d行第%d列数据不一致'%(i,j))
                    print (row1[i][j].value,row2[i][j].value)   
                    ws1.cell(row=i+1,column=j).fill=yellowfill
                    ws2.cell(row=i+1,column=j).fill=yellowfill  
        #if not flag:
            #return
                    
    wb1.save(filepath1)
    wb2.save(filepath2)
                 

        
                    
        
if __name__=='__main__':
    xls2=r'D:\zhaocuixia\Desktop\zhaocuixia_pop其他费用日报_20190929182100.xls.xlsx'
    xls1=r'D:\zhaocuixia\Desktop\zcxxs3.xlsx'
    #diff_excel(xls1,xls2)
    diff(xls1,xls2)
    
        
